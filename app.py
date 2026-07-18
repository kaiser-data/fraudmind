#!/usr/bin/env python3
"""fraudmind local server: review console + Cognee brain proxy.

Stdlib only (no FastAPI dependency — broken starlette on this machine).
Serves the React review console (frontend/dist) when built, the legacy
brain.html dashboard, and a JSON API:

  POST /api/upload        multipart dossier upload -> runs pipeline
  POST /api/analyze       run pipeline on a server-side dossier path
  GET  /api/status        pipeline progress log
  GET  /api/findings      deterministic findings (checks.py output)
  GET  /api/explanations  LLM auditor language (explain.py output)
  GET  /api/graph         transaction points for amount/time context
  GET/POST /api/verdicts  reviewer decisions per finding
  GET  /api/report        markdown report incl. verdicts (download)
  POST /api/brain/update  write confirmed frauds back as brain nodes
  POST /api/ask           Cognee Cloud recall + findings cross-check

The browser never sees API keys; they stay in .env server-side.
Run: python3 app.py   ->  http://127.0.0.1:8600
"""
import json
import os
import re
import subprocess
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

ROOT = Path(__file__).parent
BUILD = ROOT / "build"
DIST = ROOT / "frontend" / "dist"
UPLOADS = ROOT / "uploads"
DATASET = "muster_verpackungen_2025"
PORT = 8600
PRACTICE_DOSSIER = ROOT / "dataset" / "Uebungsdaten Muster Verpackungen"

MIME = {".html": "text/html; charset=utf-8", ".js": "text/javascript",
        ".css": "text/css", ".svg": "image/svg+xml", ".png": "image/png",
        ".ico": "image/x-icon", ".json": "application/json",
        ".woff2": "font/woff2"}


def load_env():
    env = {}
    env_file = ROOT / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if "=" in line and not line.strip().startswith("#"):
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"')
    base = os.environ.get("COGNEE_BASE_URL") or env.get("COGNEE_BASE_URL")
    key = os.environ.get("COGNEE_API_KEY") or env.get("COGNEE_API_KEY")
    if not base or not key:
        raise SystemExit("COGNEE_BASE_URL / COGNEE_API_KEY missing (.env or env)")
    return base.rstrip("/"), key


COGNEE_URL, COGNEE_KEY = load_env()

# ---------------------------------------------------------------- pipeline job
JOB = {"running": False, "stage": "idle", "log": [], "error": None,
       "done": False}
JOB_LOCK = threading.Lock()


def job_log(stage, message):
    with JOB_LOCK:
        JOB["stage"] = stage
        JOB["log"].append({"stage": stage, "message": message,
                           "at": time.strftime("%H:%M:%S")})


def run_pipeline(dossier: Path):
    steps = [
        ("ingest", [sys.executable, "ingest.py", str(dossier)],
         "Parsing GDPdU dossier, linking document chains"),
        ("checks", [sys.executable, "checks.py"],
         "Running deterministic control tests"),
        ("explain", [sys.executable, "explain.py"],
         "Drafting auditor language (GPT)"),
    ]
    try:
        for stage, cmd, label in steps:
            job_log(stage, label)
            proc = subprocess.run(cmd, cwd=ROOT, capture_output=True,
                                  text=True, timeout=600)
            if proc.returncode != 0:
                if stage == "explain":  # LLM layer is best-effort (quota)
                    job_log(stage, "AI explanations unavailable "
                                   "(engine findings unaffected)")
                    continue
                raise RuntimeError(
                    f"{stage} failed: {(proc.stderr or proc.stdout)[-400:]}")
            tail = (proc.stdout or "").strip().splitlines()
            if tail:
                job_log(stage, tail[-1][:200])
        job_log("done", "Case brain ready for review")
        with JOB_LOCK:
            JOB["done"] = True
    except Exception as e:
        with JOB_LOCK:
            JOB["error"] = f"{type(e).__name__}: {e}"
            JOB["log"].append({"stage": "error", "message": JOB["error"],
                               "at": time.strftime("%H:%M:%S")})
    finally:
        with JOB_LOCK:
            JOB["running"] = False


def start_pipeline(dossier: Path):
    with JOB_LOCK:
        if JOB["running"]:
            return False
        JOB.update({"running": True, "stage": "start", "log": [],
                    "error": None, "done": False})
    threading.Thread(target=run_pipeline, args=(dossier,),
                     daemon=True).start()
    return True


# ---------------------------------------------------------------- upload
SAFE_PART = re.compile(r"[^\w\-. äöüÄÖÜß]")


def safe_relpath(name: str) -> Path | None:
    parts = []
    for part in Path(name.replace("\\", "/")).parts:
        if part in ("..", "/", ""):
            continue
        parts.append(SAFE_PART.sub("_", part))
    return Path(*parts) if parts else None


def parse_multipart(headers, body: bytes):
    """Minimal multipart/form-data parser: yields (filename, payload)."""
    ctype = headers.get("Content-Type", "")
    match = re.search(r'boundary="?([^";]+)"?', ctype)
    if not match:
        return
    boundary = b"--" + match.group(1).encode()
    for part in body.split(boundary):
        if b"\r\n\r\n" not in part:
            continue
        head, _, payload = part.partition(b"\r\n\r\n")
        fn = re.search(rb'filename="([^"]*)"', head)
        if not fn or not fn.group(1):
            continue
        payload = payload.rstrip(b"\r\n").removesuffix(b"--").rstrip(b"\r\n")
        yield fn.group(1).decode("utf-8", errors="replace"), payload


# ---------------------------------------------------------------- source viewer
ACTIVE = {"dossier": PRACTICE_DOSSIER}
SRC_REF = re.compile(
    r"^(?P<path>[^:#]+?)(?:#(?P<sheet>[^:]+))?(?::(?P<kind>row|page)(?P<n>\d+))?$")
CONTEXT_ROWS = 3


def _index_columns(dossier: Path, module: str, filename: str) -> list[str]:
    """Column names for a GDPdU txt table from the module's index.xml."""
    idx_file = dossier / module / "index.xml"
    if not idx_file.exists():
        return []
    idx = idx_file.read_text(encoding="utf-8", errors="replace")
    for table in re.findall(r"<Table>(.*?)</Table>", idx, re.S):
        url = re.search(r"<URL>([^<]+)</URL>", table)
        if url and url.group(1) == filename:
            return re.findall(r"<Name>([^<]+)</Name>", table)[1:]
    return []


def source_view(ref: str) -> dict:
    """Resolve a provenance ref (file:rowN / file#Sheet:rowN / file:pageN)
    to the source content around the cited location."""
    m = SRC_REF.match(ref.strip())
    if not m:
        return {"error": "unparseable reference"}
    dossier = ACTIVE["dossier"]
    target = (dossier / m.group("path")).resolve()
    if dossier.resolve() not in target.parents or not target.is_file():
        return {"error": f"source file not in current dossier: {m.group('path')}"}
    n = int(m.group("n")) if m.group("n") else None
    suffix = target.suffix.lower()

    if suffix == ".pdf" and n:
        try:
            txt = subprocess.run(
                ["pdftotext", "-layout", "-f", str(n), "-l", str(n),
                 str(target), "-"],
                capture_output=True, text=True, timeout=30).stdout
        except Exception as e:
            return {"error": f"pdf extraction failed: {e}"}
        return {"ref": ref, "kind": "pdf", "page": n,
                "text": txt.strip()[:6000]}

    if suffix == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(target, data_only=True, read_only=True)
        sheet = m.group("sheet")
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
        data = [[("" if c is None else str(c)) for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        header_i = next((i for i, r in enumerate(data)
                         if sum(bool(c) for c in r) >= 2), 0)
        header = data[header_i] if data else []
        if n is None:
            n = header_i + 2
        lo = max(header_i + 1, n - 1 - CONTEXT_ROWS)
        rows = [{"n": i + 1, "cells": data[i], "target": i + 1 == n}
                for i in range(lo, min(len(data), n - 1 + CONTEXT_ROWS + 1))]
        return {"ref": ref, "kind": "table", "file": m.group("path"),
                "sheet": ws.title, "header": header, "rows": rows}

    if suffix in (".csv", ".txt"):
        lines = target.read_text(encoding="latin-1",
                                 errors="replace").splitlines()
        parts = Path(m.group("path")).parts
        if suffix == ".txt" and len(parts) == 2:
            header = _index_columns(dossier, parts[0], parts[1])
        else:
            header = lines[0].split(";") if lines else []
        if n is None:
            n = 2 if suffix == ".csv" else 1
        lo = max(2 if suffix == ".csv" else 1, n - CONTEXT_ROWS)
        rows = [{"n": i, "cells": lines[i - 1].split(";"), "target": i == n}
                for i in range(lo, min(len(lines), n + CONTEXT_ROWS) + 1)]
        return {"ref": ref, "kind": "table", "file": m.group("path"),
                "header": [h.strip('"') for h in header], "rows": rows}

    return {"error": f"no viewer for {suffix or 'this file type'}"}


# ---------------------------------------------------------------- graph data
GRAPH_CACHE = {"mtime": None, "points": None}


def parse_amount(raw) -> float | None:
    if raw in (None, ""):
        return None
    try:
        return float(str(raw).replace(".", "").replace(",", "."))
    except ValueError:
        return None


def graph_points():
    src = BUILD / "chains.json"
    if not src.exists():
        return []
    mtime = src.stat().st_mtime
    if GRAPH_CACHE["mtime"] == mtime:
        return GRAPH_CACHE["points"]
    points = []
    for cid, chain in json.loads(src.read_text()).items():
        if not cid:
            continue
        kind = cid[:2]
        date = amount = party = konto = prov = None
        fak = chain.get("faktura")
        sub = chain.get("subledger") or []
        goods = chain.get("goods") or []
        if fak:
            date = fak.get("FAKTURADATUM")
            amount = parse_amount(fak.get("BETRAG_EUR"))
            party, konto = fak.get("DEBITORNAME"), fak.get("DEBITOR")
            prov = fak.get("_prov")
        elif sub:
            row = sub[0]
            date = row.get("BUCHUNGSDATUM")
            amount = parse_amount(row.get("BUCHUNGSBETRAG"))
            if amount is not None:
                amount = abs(amount)
            konto = row.get("LIEFERANTENKONTONUMMER")
            party = row.get("BUCHUNGSTEXT")
            prov = row.get("_prov")
        elif goods:
            row = goods[0]
            date = (row.get("WARENEINGANG_DATUM")
                    or row.get("WARENAUSGANG_DATUM"))
            amount = parse_amount(row.get("BETRAG_EUR"))
            party = row.get("KREDITORNAME") or row.get("DEBITORNAME")
            konto = row.get("KREDITOR") or row.get("DEBITOR")
            prov = row.get("_prov")
        if not date or amount is None:
            continue
        points.append({"id": cid, "kind": kind, "date": date,
                       "amount": amount, "party": party or "",
                       "konto": konto or "", "prov": prov or ""})
    GRAPH_CACHE.update({"mtime": mtime, "points": points})
    return points


# ---------------------------------------------------------------- json stores
def read_json(path: Path, default):
    if path.exists():
        return json.loads(path.read_text())
    return default


def clean_explanations():
    raw = read_json(BUILD / "explanations.json", {})
    return {fid: exp for fid, exp in raw.items() if "error" not in exp}


def read_verdicts():
    return read_json(BUILD / "verdicts.json", {})


def write_verdict(fid, verdict, note):
    verdicts = read_verdicts()
    verdicts[fid] = {"verdict": verdict, "note": note,
                     "at": time.strftime("%Y-%m-%d %H:%M:%S")}
    (BUILD / "verdicts.json").write_text(
        json.dumps(verdicts, indent=2, ensure_ascii=False))
    return verdicts


# ---------------------------------------------------------------- report
VERDICT_LABEL = {"confirmed": "CONFIRMED FRAUD", "followup": "FOLLOW-UP",
                 "dismissed": "DISMISSED"}


def build_report() -> str:
    findings = read_json(BUILD / "findings.json", [])
    explanations = clean_explanations()
    verdicts = read_verdicts()
    confirmed = [f for f in findings
                 if verdicts.get(f["id"], {}).get("verdict") == "confirmed"]
    total_confirmed = sum(f.get("amount_eur") or 0 for f in confirmed)
    lines = [
        "# fraudmind — Fraud Findings Report",
        "",
        f"Generated {time.strftime('%Y-%m-%d %H:%M')} · dataset {DATASET}",
        "",
        f"- Findings raised by engine: **{len(findings)}**",
        f"- Reviewed by auditor: **{len(verdicts)} / {len(findings)}**",
        f"- Confirmed fraud: **{len(confirmed)}** "
        f"(quantified impact **{total_confirmed:,.2f} EUR**)",
        "",
        "Every figure below is produced by the deterministic control engine "
        "and cites its source document. AI text is language only — it never "
        "decides numbers. Verdicts are the human reviewer's.",
        "",
    ]
    for f in findings:
        v = verdicts.get(f["id"])
        exp = explanations.get(f["id"], {})
        lines.append(f"## {f['id']} — {f['title']}")
        meta = (f"**{f['tier']}** · severity {f['severity']} · "
                f"confidence {f.get('confidence', 0):.0%}")
        if f.get("amount_eur"):
            meta += f" · {f['amount_eur']:,.2f} EUR"
        lines.append(meta)
        if v:
            lines.append(f"**Reviewer verdict: "
                         f"{VERDICT_LABEL.get(v['verdict'], v['verdict'])}**"
                         + (f" — {v['note']}" if v.get("note") else "")
                         + f" ({v['at']})")
        else:
            lines.append("_Not yet reviewed._")
        lines.append("")
        lines.append(f.get("explanation", ""))
        if exp.get("explanation_de"):
            lines.append("")
            lines.append(f"> DE: {exp['explanation_de']}")
        if exp.get("recommended_action_en"):
            lines.append(f"> Next steps: {exp['recommended_action_en']}")
        lines.append("")
        lines.append("Provenance: " + "; ".join(
            f"`{p}`" for p in f.get("provenance", [])))
        lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------- brain nodes
def brain_update():
    """Write confirmed frauds back into the case brain as fraud nodes."""
    findings = read_json(BUILD / "findings.json", [])
    verdicts = read_verdicts()
    nodes = []
    for f in findings:
        v = verdicts.get(f["id"], {}).get("verdict")
        if v not in ("confirmed", "followup"):
            continue
        nodes.append({
            "node": f"FRAUD_{f['id']}",
            "status": "confirmed_fraud" if v == "confirmed" else "follow_up",
            "title": f["title"],
            "amount_eur": f.get("amount_eur"),
            "summary": f.get("explanation", ""),
            "provenance": f.get("provenance", []),
        })
    (BUILD / "brain_fraud_nodes.json").write_text(
        json.dumps(nodes, indent=2, ensure_ascii=False))
    pushed = False
    try:  # best-effort push into Cognee so ask-the-brain knows the verdicts
        texts = [f"Fraud node {n['node']} ({n['status']}): {n['title']}. "
                 f"Amount {n['amount_eur']} EUR. {n['summary']} "
                 f"Sources: {'; '.join(n['provenance'][:4])}" for n in nodes]
        if texts:
            req = urllib.request.Request(
                f"{COGNEE_URL}/api/v1/add",
                data=json.dumps({"data": texts,
                                 "datasetName": DATASET}).encode(),
                headers={"X-Api-Key": COGNEE_KEY,
                         "Content-Type": "application/json"},
                method="POST")
            with urllib.request.urlopen(req, timeout=60):
                pushed = True
    except Exception:
        pushed = False
    return {"nodes": len(nodes), "cognee_pushed": pushed}


# ---------------------------------------------------------------- cognee ask
def cognee_recall(question):
    req = urllib.request.Request(
        f"{COGNEE_URL}/api/v1/recall",
        data=json.dumps({"query": question, "datasets": [DATASET]}).encode(),
        headers={"X-Api-Key": COGNEE_KEY, "Content-Type": "application/json"},
        method="POST")
    with urllib.request.urlopen(req, timeout=90) as r:
        return json.loads(r.read().decode())


def related_findings(question):
    words = {w.lower().strip("?.,") for w in question.split() if len(w) > 3}
    hits = []
    for f in read_json(BUILD / "findings.json", []):
        hay = (f["title"] + " " + f["explanation"]).lower()
        if sum(1 for w in words if w in hay) >= 2:
            hits.append({"id": f["id"], "tier": f["tier"], "title": f["title"],
                         "provenance": f["provenance"][:4]})
    return hits[:5]


# ---------------------------------------------------------------- http
class Handler(BaseHTTPRequestHandler):
    def _send(self, code, body, ctype="application/json", extra=None):
        data = body if isinstance(body, bytes) else json.dumps(body).encode()
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        for k, v in (extra or {}).items():
            self.send_header(k, v)
        self.end_headers()
        self.wfile.write(data)

    def _static(self, rel):
        path = (DIST / rel).resolve()
        if DIST.resolve() not in path.parents or not path.is_file():
            self._send(404, {"detail": "not found"})
            return
        self._send(200, path.read_bytes(),
                   MIME.get(path.suffix, "application/octet-stream"))

    def do_GET(self):
        route = self.path.split("?")[0]
        if route == "/" and (DIST / "index.html").exists():
            self._send(200, (DIST / "index.html").read_bytes(), MIME[".html"])
        elif route in ("/", "/brain.html"):
            self._send(200, (ROOT / "brain.html").read_bytes(), MIME[".html"])
        elif route.startswith("/assets/"):
            self._static(route.lstrip("/"))
        elif route == "/api/findings":
            self._send(200, read_json(BUILD / "findings.json", []))
        elif route == "/api/explanations":
            self._send(200, clean_explanations())
        elif route == "/api/graph":
            self._send(200, graph_points())
        elif route == "/api/verdicts":
            self._send(200, read_verdicts())
        elif route == "/api/status":
            with JOB_LOCK:
                self._send(200, dict(JOB))
        elif route == "/api/report":
            self._send(200, build_report().encode(),
                       "text/markdown; charset=utf-8",
                       {"Content-Disposition":
                        'attachment; filename="fraudmind_report.md"'})
        elif route == "/api/source":
            query = urllib.parse.parse_qs(
                urllib.parse.urlsplit(self.path).query)
            ref = (query.get("ref") or [""])[0]
            result = source_view(ref) if ref else {"error": "ref required"}
            self._send(200 if "error" not in result else 404, result)
        else:
            self._send(404, {"detail": "not found"})

    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(length)

    def do_POST(self):
        try:
            route = self.path.split("?")[0]
            if route == "/api/upload":
                self._post_upload()
            elif route == "/api/analyze":
                body = json.loads(self._read_body() or b"{}")
                dossier = Path(body.get("path") or PRACTICE_DOSSIER)
                if not dossier.is_dir():
                    self._send(400, {"detail": f"no dossier at {dossier}"})
                    return
                ACTIVE["dossier"] = dossier
                started = start_pipeline(dossier)
                self._send(200 if started else 409,
                           {"started": started})
            elif route == "/api/verdicts":
                body = json.loads(self._read_body() or b"{}")
                fid = str(body.get("id", "")).strip()
                verdict = str(body.get("verdict", "")).strip()
                if not fid or verdict not in ("confirmed", "followup",
                                              "dismissed"):
                    self._send(400, {"detail": "need id + verdict "
                                     "(confirmed|followup|dismissed)"})
                    return
                self._send(200, write_verdict(
                    fid, verdict, str(body.get("note", ""))[:500]))
            elif route == "/api/brain/update":
                self._send(200, brain_update())
            elif route == "/api/ask":
                self._post_ask()
            else:
                self._send(404, {"detail": "not found"})
        except urllib.error.URLError as e:
            self._send(502, {"detail": f"upstream unreachable: {e}"})
        except Exception as e:  # surface any handler error to the UI
            self._send(500, {"detail": f"{type(e).__name__}: {e}"})

    def _post_upload(self):
        body = self._read_body()
        case_dir = UPLOADS / time.strftime("case_%Y%m%d_%H%M%S")
        saved = 0
        for filename, payload in parse_multipart(self.headers, body):
            rel = safe_relpath(filename)
            if rel is None:
                continue
            dest = case_dir / rel
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(payload)
            saved += 1
        if not saved:
            self._send(400, {"detail": "no files received"})
            return
        ACTIVE["dossier"] = case_dir
        started = start_pipeline(case_dir)
        self._send(200, {"saved": saved, "case": case_dir.name,
                         "started": started})

    def _post_ask(self):
        body = json.loads(self._read_body() or b"{}")
        q = str(body.get("question", "")).strip()
        if not q:
            self._send(400, {"detail": "empty question"})
            return
        results = cognee_recall(q)
        answers = [{"text": item.get("text", ""),
                    "source": item.get("source", ""),
                    "dataset": item.get("dataset_name", "")}
                   for item in results if item.get("text")]
        self._send(200, {"answers": answers,
                         "related_findings": related_findings(q)})

    def log_message(self, fmt, *args):
        print(f"{self.address_string()} {fmt % args}")


if __name__ == "__main__":
    print(f"fraudmind on http://127.0.0.1:{PORT}  (Cognee: {COGNEE_URL})")
    ThreadingHTTPServer(("127.0.0.1", PORT), Handler).serve_forever()
